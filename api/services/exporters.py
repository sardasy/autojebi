"""Markdown 표 초안 → 실파일 변환 (M11 — 서류 자동화 v2).

- Excel: ``openpyxl``로 직접 생성. 빠르고 외부 의존 없음.
- HWP: ``milim-hwp-agent``의 ``/document/insert-table`` 엔드포인트 호출. 미구현 시 502 fallback.

라우터에서 호출되는 순수 함수. 디스크 경로는 ``export_dir()``에서 결정.
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import tempfile
import uuid
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from fastapi import HTTPException

from api.config import settings
from api.models.notices import ExportRecord
from api.services.hwp_agent_client import HwpAgentClient, HwpAgentError

log = logging.getLogger(__name__)

_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_.\-]")

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HWP_MIME = "application/x-hwp"
COMPLIANCE_EXCEL_V1 = "compliance_excel_v1"
COMPLIANCE_EXCEL_V2 = "compliance_excel_v2"


class ExportError(HTTPException):
    """HTTPException 서브클래스 — 라우터에서 raise 그대로."""


def export_dir() -> Path:
    root = settings.export_dir.strip() or os.path.join(
        tempfile.gettempdir(), "autojebi", "exports"
    )
    path = Path(root)
    path.mkdir(parents=True, exist_ok=True)
    return path


def notice_dir(notice_no: str) -> Path:
    safe = _SAFE_NAME_RE.sub("_", notice_no.strip()) or "unknown"
    path = export_dir() / safe
    path.mkdir(parents=True, exist_ok=True)
    return path


def parse_markdown_table(markdown: str) -> tuple[list[str], list[list[str]]]:
    """``| h1 | h2 |`` Markdown 표를 (headers, rows)로 변환.

    헤더-구분선-데이터 3행 이상 가정. 구분선(`| --- |`)은 자동 스킵.
    표 형식이 아니면 단일 컬럼으로 fallback.
    """
    lines = [line.strip() for line in markdown.splitlines() if line.strip()]
    if not lines:
        return ["내용"], []

    if all(line.startswith("|") for line in lines):
        rows: list[list[str]] = []
        for line in lines:
            cells = [c.strip() for c in line.strip("|").split("|")]
            rows.append(cells)
        if len(rows) >= 2 and all(set(c) <= set("-: ") for c in rows[1]):
            headers = rows[0]
            data = rows[2:]
        else:
            headers = rows[0]
            data = rows[1:]
        if not headers:
            headers = ["내용"]
        return headers, data

    # 단일 컬럼 fallback — 한 줄 한 행
    return ["내용"], [[line] for line in lines]


def get_technical_compliance_draft(document_automation: dict) -> dict[str, Any]:
    drafts = document_automation.get("drafts") if isinstance(document_automation, dict) else None
    if not isinstance(drafts, dict):
        raise ExportError(status_code=409, detail="no drafts available")
    draft = drafts.get("technical_compliance")
    if not isinstance(draft, dict):
        raise ExportError(
            status_code=409,
            detail="technical_compliance draft missing — run /documents/analyze first",
        )
    return draft


def build_excel(
    *,
    notice_no: str,
    draft: dict[str, Any],
    title: str | None = None,
    version: str = COMPLIANCE_EXCEL_V2,
    notice_meta: dict[str, Any] | None = None,
    validation_warnings: list[dict[str, Any]] | None = None,
) -> ExportRecord:
    """``technical_compliance`` 초안을 .xlsx로 저장.

    v1은 기존 단일 시트 포맷이고, v2는 검토용 메타/경고 시트와 필터를 추가한다.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    content = str(draft.get("content") or "")
    if not content:
        raise ExportError(status_code=409, detail="draft has empty content")
    headers, rows = parse_markdown_table(content)

    wb = Workbook()
    ws = wb.active
    ws.title = "규격대응표"
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = bold
        if version == COMPLIANCE_EXCEL_V2:
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if version == COMPLIANCE_EXCEL_V2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col_index, _ in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = (
            28 if version == COMPLIANCE_EXCEL_V2 else 22
        )

    if title:
        ws.cell(row=ws.max_row + 2, column=1, value=title)

    if version == COMPLIANCE_EXCEL_V2:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        meta_ws = wb.create_sheet("검토요약")
        meta_ws.append(["항목", "값"])
        meta_ws["A1"].font = bold
        meta_ws["B1"].font = bold
        meta = {
            "notice_no": notice_no,
            "title": title or "",
            "version": version,
            "generated_at": datetime.now(tz=UTC).isoformat(),
            **(notice_meta or {}),
        }
        for key, value in meta.items():
            meta_ws.append([key, str(value or "")])
        warnings = validation_warnings or []
        if warnings:
            meta_ws.append(["warnings", f"{len(warnings)}건"])
            for item in warnings:
                meta_ws.append([str(item.get("stage") or "warning"), str(item.get("detail") or item)])
        meta_ws.column_dimensions["A"].width = 24
        meta_ws.column_dimensions["B"].width = 80

    target = notice_dir(notice_no) / f"compliance_{uuid.uuid4().hex}.xlsx"
    try:
        wb.save(str(target))
    except OSError as exc:
        log.exception("[exporters] excel write failed: %s", exc)
        raise ExportError(status_code=500, detail=f"failed to write excel: {exc}") from exc

    return with_file_metadata(
        ExportRecord(
            kind="excel",
            draft_id="technical_compliance",
            output_path=str(target),
            mime=EXCEL_MIME,
            generated_at=datetime.now(tz=UTC).isoformat(),
            notes=str(draft.get("label") or ""),
            version=version,
            template_version=version,
            validation_status="warning" if validation_warnings else "passed",
            validation_errors=validation_warnings or [],
        )
    )


def build_hwp(
    *,
    client: HwpAgentClient,
    notice_no: str,
    draft: dict[str, Any],
    title: str | None = None,
) -> ExportRecord:
    """``technical_compliance`` 초안을 milim-hwp-agent로 위임해 .hwp 생성.

    에이전트에 ``/document/insert-table``가 없으면 502가 raise됨 — 라우터에서 그대로 전달.
    """
    content = str(draft.get("content") or "")
    if not content:
        raise ExportError(status_code=409, detail="draft has empty content")
    headers, rows = parse_markdown_table(content)
    output_path = str(notice_dir(notice_no) / f"compliance_{uuid.uuid4().hex}.hwp")

    try:
        result = client.generate_compliance_table(
            output_path=output_path,
            title=title or str(draft.get("label") or "규격대응표"),
            headers=headers,
            rows=rows,
        )
    except HwpAgentError as exc:
        raise ExportError(status_code=502, detail=f"hwp agent failed: {exc}") from exc

    resolved_path = str(result.get("output_path") or output_path)
    return with_file_metadata(
        ExportRecord(
            kind="hwp",
            draft_id="technical_compliance",
            output_path=resolved_path,
            mime=HWP_MIME,
            generated_at=datetime.now(tz=UTC).isoformat(),
            notes=str(result.get("sheet_count") or "") or None,
            version="technical_compliance_hwp_v1",
            template_version="document_insert_table_v1",
            validation_status="passed",
        )
    )


def with_file_metadata(export: ExportRecord) -> ExportRecord:
    path = Path(export.output_path)
    if not path.is_file():
        return export
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return export.model_copy(
            update={"file_size": path.stat().st_size, "sha256": digest.hexdigest()}
        )
    except OSError:
        return export


def merge_export_into_document_automation(
    document_automation: dict,
    export: ExportRecord,
) -> dict:
    """exports 배열에 추가 — 같은 (kind, draft_id) 이전 기록은 교체."""
    updated = dict(document_automation)
    exports = list(updated.get("exports") or [])
    survivors = [
        item
        for item in exports
        if not (
            isinstance(item, dict)
            and item.get("kind") == export.kind
            and item.get("draft_id") == export.draft_id
        )
    ]
    survivors.append(export.model_dump())
    updated["exports"] = survivors
    return updated


def lookup_export(
    document_automation: dict,
    *,
    kind: str,
    draft_id: str = "technical_compliance",
) -> dict | None:
    """이미 생성된 export 메타데이터를 찾는다 (download 라우터용)."""
    exports = document_automation.get("exports") if isinstance(document_automation, dict) else None
    if not isinstance(exports, list):
        return None
    for item in exports:
        if (
            isinstance(item, dict)
            and item.get("kind") == kind
            and item.get("draft_id") == draft_id
        ):
            return item
    return None
