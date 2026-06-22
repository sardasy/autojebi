"""POST/GET /notices/{notice_no}/documents/exports/{kind}* 통합 테스트.

Excel은 실제 디스크에 쓰고, HWP는 HwpAgentClient mock으로 격리.
"""

from __future__ import annotations

from datetime import UTC, datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, insert, select
from sqlalchemy.pool import StaticPool

from api.config import settings
from api.main import app
from api.routers import notices as notices_router
from api.routers.notices import bid_pipeline, metadata, notice_exports, notice_spec_items
from api.services.hwp_agent_client import HwpAgentError


@pytest.fixture
def sqlite_engine(tmp_path, monkeypatch):
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    metadata.create_all(engine)
    monkeypatch.setattr("api.db.get_engine", lambda: engine)
    monkeypatch.setattr(settings, "api_key", "")
    monkeypatch.setattr(settings, "anthropic_api_key", "")
    monkeypatch.setattr(settings, "export_dir", str(tmp_path / "exports"))
    return engine


@pytest.fixture
def client(sqlite_engine, monkeypatch):
    monkeypatch.setattr(settings, "scheduler_enabled", False)
    return TestClient(app)


_MD = (
    "| 항목 | 공고 요구사양 | 추출/추천 값 | 확인 |\n"
    "| --- | --- | --- | --- |\n"
    "| 품목 | 공고문 확인 | 변압기 | 담당자 검토 |\n"
)


def _seed(engine, *, with_drafts=True):
    now = datetime.now(tz=UTC)
    drafts = (
        {"technical_compliance": {"kind": "markdown", "label": "규격대응표", "content": _MD}}
        if with_drafts
        else {}
    )
    with engine.begin() as conn:
        conn.execute(
            insert(bid_pipeline).values(
                notice_no="DOC-1",
                title="ABB 변압기 구매",
                source="G2B",
                raw={},
                category="ABB장비",
                fit_score=80,
                assignee="이용문",
                analysis={
                    "document_automation": {
                        "checklist": [],
                        "drafts": drafts,
                        "risks": [],
                        "generated_at": now.isoformat(),
                        "source": "rule",
                        "ready_for_submission": True,
                        "missing_required": [],
                        "errors": [],
                        "uploads": [],
                        "exports": [],
                    }
                },
                status="analyzed",
                created_at=now,
                updated_at=now,
            )
        )


def _seed_spec_item(engine):
    with engine.begin() as conn:
        conn.execute(
            insert(notice_spec_items).values(
                notice_no="DOC-1",
                item_key="product_category",
                label="품목",
                required_value="변압기",
                proposed_value="ABB 변압기",
                unit="",
                category="technical",
                source="test",
                confidence=0.95,
                evidence={},
                status="matched",
                sort_order=1,
            )
        )


def test_export_excel_writes_xlsx_and_persists(client, sqlite_engine):
    _seed(sqlite_engine)
    r = client.post("/notices/DOC-1/documents/exports/excel")
    assert r.status_code == 200, r.text
    body = r.json()["export"]
    assert body["id"]
    assert body["kind"] == "excel"
    assert body["version"] == "compliance_excel_v2"
    assert body["validation_status"] == "passed"
    assert body["file_size"] > 0
    assert body["sha256"]
    assert body["output_path"].endswith(".xlsx")

    wb = load_workbook(body["output_path"])
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "항목"
    assert ws.cell(row=2, column=1).value == "품목"
    assert "검토요약" in wb.sheetnames
    with sqlite_engine.begin() as conn:
        row = conn.execute(select(notice_exports).where(notice_exports.c.notice_no == "DOC-1")).mappings().one()
    assert row["kind"] == "excel"
    assert row["version"] == "compliance_excel_v2"
    assert row["file_size"] == body["file_size"]
    assert row["deleted_at"] is None


def test_export_excel_v1_keeps_legacy_single_sheet(client, sqlite_engine):
    _seed(sqlite_engine)
    r = client.post(
        "/notices/DOC-1/documents/exports/excel",
        json={"version": "compliance_excel_v1"},
    )
    assert r.status_code == 200, r.text
    body = r.json()["export"]
    assert body["version"] == "compliance_excel_v1"
    wb = load_workbook(body["output_path"])
    assert wb.sheetnames == ["규격대응표"]


def test_export_excel_409_when_draft_missing(client, sqlite_engine):
    _seed(sqlite_engine, with_drafts=False)
    r = client.post("/notices/DOC-1/documents/exports/excel")
    assert r.status_code == 409


def test_export_hwp_returns_502_when_agent_errors(client, sqlite_engine, monkeypatch):
    _seed(sqlite_engine)
    _seed_spec_item(sqlite_engine)

    class FakeClient:
        def generate_compliance_table(self, **kw):
            raise HwpAgentError("agent missing /document/insert-table")

    monkeypatch.setattr(notices_router, "_make_hwp_agent_client", lambda: FakeClient())
    r = client.post("/notices/DOC-1/documents/exports/hwp")
    assert r.status_code == 502
    assert "hwp agent failed" in r.json()["detail"]


def test_export_hwp_uses_agent_output_path_and_persists(client, sqlite_engine, monkeypatch, tmp_path):
    _seed(sqlite_engine)
    _seed_spec_item(sqlite_engine)
    final_path = tmp_path / "compliance.hwp"
    final_path.write_bytes(b"HWPCONTENT")

    class FakeClient:
        def generate_compliance_table(self, **kw):
            return {"output_path": str(final_path), "sheet_count": 1}

    monkeypatch.setattr(notices_router, "_make_hwp_agent_client", lambda: FakeClient())
    r = client.post("/notices/DOC-1/documents/exports/hwp")
    assert r.status_code == 200, r.text
    assert r.json()["export"]["output_path"] == str(final_path)


def test_export_download_streams_generated_file(client, sqlite_engine):
    _seed(sqlite_engine)
    create = client.post("/notices/DOC-1/documents/exports/excel")
    r = client.get("/notices/DOC-1/documents/exports/excel/download")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # xlsx zip 매직넘버 = PK\x03\x04
    assert r.content[:2] == b"PK"

    export_id = create.json()["export"]["id"]
    r2 = client.get(f"/notices/DOC-1/documents/exports/by-id/{export_id}/download")
    assert r2.status_code == 200
    assert r2.content[:2] == b"PK"


def test_export_regeneration_soft_deletes_previous_active_row(client, sqlite_engine):
    _seed(sqlite_engine)
    client.post("/notices/DOC-1/documents/exports/excel")
    client.post("/notices/DOC-1/documents/exports/excel")

    with sqlite_engine.begin() as conn:
        total = conn.execute(select(func.count()).select_from(notice_exports)).scalar_one()
        active = conn.execute(
            select(func.count())
            .select_from(notice_exports)
            .where(notice_exports.c.deleted_at.is_(None))
        ).scalar_one()
    assert total == 2
    assert active == 1


def test_export_download_404_when_not_generated(client, sqlite_engine):
    _seed(sqlite_engine)
    r = client.get("/notices/DOC-1/documents/exports/excel/download")
    assert r.status_code == 404


def test_export_hwp_precompose_validation_blocks_missing_required_doc(client, sqlite_engine):
    _seed(sqlite_engine)
    with sqlite_engine.begin() as conn:
        analysis = conn.execute(
            select(bid_pipeline.c.analysis).where(bid_pipeline.c.notice_no == "DOC-1")
        ).scalar_one()
        docs = analysis["document_automation"]
        docs["checklist"] = [
            {
                "id": "business_registration",
                "name": "사업자등록증",
                "required": True,
                "status": "needed",
            }
        ]
        conn.execute(
            bid_pipeline.update()
            .where(bid_pipeline.c.notice_no == "DOC-1")
            .values(analysis=analysis)
        )

    r = client.post("/notices/DOC-1/documents/exports/hwp")

    assert r.status_code == 409
    detail = r.json()["detail"]
    assert detail["errors"][0]["stage"] == "pre_compose.spec_items"
    with sqlite_engine.begin() as conn:
        rows = conn.execute(select(notice_exports)).all()
    assert rows == []
