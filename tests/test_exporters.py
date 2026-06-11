"""api/services/exporters.py 단위 테스트.

Excel 파일 생성은 실제로 디스크에 쓰고 openpyxl로 다시 읽어서 검증.
HWP는 HwpAgentClient mock으로 격리.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from api.config import settings
from api.services.exporters import (
    ExportError,
    build_excel,
    build_hwp,
    get_technical_compliance_draft,
    lookup_export,
    merge_export_into_document_automation,
    parse_markdown_table,
)
from api.services.hwp_agent_client import HwpAgentError


@pytest.fixture(autouse=True)
def _isolate_export_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "export_dir", str(tmp_path))


_MD_TABLE = (
    "| 항목 | 공고 요구사양 | 추출/추천 값 | 확인 |\n"
    "| --- | --- | --- | --- |\n"
    "| 품목 | 공고문 확인 | 변압기 | 담당자 검토 |\n"
    "| 정격용량 | 공고문 확인 | 1000kVA | 담당자 검토 |\n"
)


def test_parse_markdown_table_strips_separator_row():
    headers, rows = parse_markdown_table(_MD_TABLE)
    assert headers == ["항목", "공고 요구사양", "추출/추천 값", "확인"]
    assert len(rows) == 2
    assert rows[0][0] == "품목"


def test_parse_markdown_table_handles_non_table_fallback():
    headers, rows = parse_markdown_table("just one line\nanother")
    assert headers == ["내용"]
    assert rows == [["just one line"], ["another"]]


def test_get_technical_compliance_draft_raises_when_missing():
    with pytest.raises(ExportError) as e:
        get_technical_compliance_draft({"drafts": {}})
    assert e.value.status_code == 409


def test_build_excel_writes_valid_xlsx_with_headers_and_rows():
    draft = {"kind": "markdown", "label": "규격대응표", "content": _MD_TABLE}
    rec = build_excel(notice_no="DOC-1", draft=draft, title="ABB 변압기")
    assert rec.kind == "excel"
    assert rec.output_path.endswith(".xlsx")
    wb = load_workbook(rec.output_path)
    ws = wb.active
    assert ws.title == "규격대응표"
    assert ws.cell(row=1, column=1).value == "항목"
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=2, column=1).value == "품목"
    assert ws.cell(row=3, column=3).value == "1000kVA"


def test_build_excel_empty_content_raises_409():
    with pytest.raises(ExportError) as e:
        build_excel(notice_no="DOC-1", draft={"kind": "markdown", "content": ""}, title=None)
    assert e.value.status_code == 409


def test_build_hwp_wraps_agent_error_as_502(monkeypatch):
    class FakeClient:
        def generate_compliance_table(self, **kw):
            raise HwpAgentError("agent has no /document/insert-table")

    draft = {"kind": "markdown", "content": _MD_TABLE}
    with pytest.raises(ExportError) as e:
        build_hwp(client=FakeClient(), notice_no="DOC-1", draft=draft, title="t")
    assert e.value.status_code == 502
    assert "hwp agent failed" in str(e.value.detail)


def test_build_hwp_uses_agent_returned_output_path(tmp_path):
    final = str(tmp_path / "resolved.hwp")

    class FakeClient:
        def generate_compliance_table(self, **kw):
            return {"output_path": final, "sheet_count": 1}

    draft = {"kind": "markdown", "content": _MD_TABLE}
    rec = build_hwp(client=FakeClient(), notice_no="DOC-1", draft=draft, title="t")
    assert rec.kind == "hwp"
    assert rec.output_path == final


def test_merge_export_replaces_same_kind_record():
    docs = {"exports": [
        {"kind": "excel", "draft_id": "technical_compliance", "output_path": "/old.xlsx",
         "mime": "x", "generated_at": "2026-01-01T00:00:00+00:00"}
    ]}
    new_rec = build_excel(
        notice_no="DOC-1",
        draft={"kind": "markdown", "content": _MD_TABLE},
        title=None,
    )
    updated = merge_export_into_document_automation(docs, new_rec)
    excels = [e for e in updated["exports"] if e["kind"] == "excel"]
    assert len(excels) == 1
    assert excels[0]["output_path"] != "/old.xlsx"


def test_lookup_export_finds_match():
    docs = {"exports": [
        {"kind": "excel", "draft_id": "technical_compliance", "output_path": "/a.xlsx",
         "mime": "x", "generated_at": "2026-01-01T00:00:00+00:00"}
    ]}
    assert lookup_export(docs, kind="excel")["output_path"] == "/a.xlsx"
    assert lookup_export(docs, kind="hwp") is None
